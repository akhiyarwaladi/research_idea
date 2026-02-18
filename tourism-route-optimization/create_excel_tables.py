#!/usr/bin/env python3
"""Create professionally formatted Excel tables for journal submission."""

import sys
import os

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "src"))

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_journal_tables():
    wb = Workbook()

    # ── Style definitions ────────────────────────────────────
    header_font = Font(name="Times New Roman", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    subheader_font = Font(name="Times New Roman", bold=True, size=10, color="1F3864")
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    body_font = Font(name="Times New Roman", size=10)
    bold_body_font = Font(name="Times New Roman", size=10, bold=True)
    title_font = Font(name="Times New Roman", bold=True, size=13, color="1F3864")
    caption_font = Font(name="Times New Roman", italic=True, size=10, color="404040")

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    header_border = Border(
        left=Side(style="thin", color="1F3864"),
        right=Side(style="thin", color="1F3864"),
        top=Side(style="medium", color="1F3864"),
        bottom=Side(style="medium", color="1F3864"),
    )

    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    best_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    def style_header_row(ws, row, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = header_border

    def style_body_cell(ws, row, col, is_alt=False, is_best=False):
        cell = ws.cell(row=row, column=col)
        cell.font = body_font
        cell.alignment = body_align
        cell.border = thin_border
        if is_best:
            cell.fill = best_fill
            cell.font = bold_body_font
        elif is_alt:
            cell.fill = alt_fill

    def auto_width(ws, extra=3):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + extra, 30)

    # ══════════════════════════════════════════════════════════
    # TABLE 1: Algorithm Performance Comparison
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Table 1 - Performance"

    ws1.merge_cells("A1:G1")
    ws1.cell(row=1, column=1, value="Table 1. Algorithm Performance Comparison (30 Independent Runs)")
    ws1.cell(row=1, column=1).font = title_font
    ws1.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    ws1.merge_cells("A2:G2")
    ws1.cell(row=2, column=1, value="Best values highlighted in green. Distance in km, time in seconds.")
    ws1.cell(row=2, column=1).font = caption_font

    headers1 = ["Algorithm", "Best (km)", "Mean (km)", "Worst (km)", "Std Dev (km)", "Mean Time (s)", "Runs"]
    for i, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=i, value=h)
    style_header_row(ws1, 4, len(headers1))

    df = pd.read_csv("results/tables/summary_table.csv")
    best_vals = {
        "Best (m)": df["Best (m)"].min(),
        "Mean (m)": df["Mean (m)"].min(),
    }
    # Exclude NN (deterministic, std=0) for Std Dev comparison
    stochastic = df[df["Algorithm"] != "NN"]
    best_std = stochastic["Std Dev (m)"].min()
    # Exclude NN for time comparison
    meta = df[df["Algorithm"].isin(["MMAS", "ACS", "GA", "SA"])]
    best_time = meta["Mean Time (s)"].min()

    for row_idx, (_, r) in enumerate(df.iterrows()):
        row = row_idx + 5
        is_alt = row_idx % 2 == 1
        data = [
            r["Algorithm"],
            round(r["Best (m)"] / 1000, 2),
            round(r["Mean (m)"] / 1000, 2),
            round(r["Worst (m)"] / 1000, 2),
            round(r["Std Dev (m)"] / 1000, 2),
            round(r["Mean Time (s)"], 4),
            int(r["Runs"]),
        ]
        for col_idx, val in enumerate(data):
            ws1.cell(row=row, column=col_idx + 1, value=val)
            is_best_cell = False
            if col_idx == 1 and r["Best (m)"] == best_vals["Best (m)"]:
                is_best_cell = True
            elif col_idx == 2 and r["Mean (m)"] == best_vals["Mean (m)"]:
                is_best_cell = True
            elif col_idx == 4 and r["Algorithm"] != "NN" and r["Std Dev (m)"] == best_std:
                is_best_cell = True
            elif col_idx == 5 and r["Algorithm"] in ("MMAS", "ACS", "GA", "SA") and r["Mean Time (s)"] == best_time:
                is_best_cell = True
            style_body_cell(ws1, row, col_idx + 1, is_alt=is_alt, is_best=is_best_cell)
        ws1.cell(row=row, column=1).alignment = left_align

    auto_width(ws1)
    ws1.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════════
    # TABLE 2: Wilcoxon Signed-Rank Tests
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Table 2 - Wilcoxon Tests")

    ws2.merge_cells("A1:F1")
    ws2.cell(row=1, column=1, value="Table 2. Wilcoxon Signed-Rank Test Results (Reference: MMAS)")
    ws2.cell(row=1, column=1).font = title_font
    ws2.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    ws2.merge_cells("A2:F2")
    ws2.cell(row=2, column=1, value="Significance level alpha = 0.05. Winner determined by lower mean distance.")
    ws2.cell(row=2, column=1).font = caption_font

    headers2 = ["Comparison", "W Statistic", "p-value", "Significant (p<0.05)", "Ref Mean (km)", "Other Mean (km)"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=i, value=h)
    style_header_row(ws2, 4, len(headers2))

    df2 = pd.read_csv("results/tables/wilcoxon_tests.csv")
    for row_idx, (_, r) in enumerate(df2.iterrows()):
        row = row_idx + 5
        is_alt = row_idx % 2 == 1
        w_stat = r["W statistic"]
        p_val = r["p-value"]
        try:
            w_display = f"{float(w_stat):.1f}"
        except (ValueError, TypeError):
            w_display = str(w_stat)
        try:
            p_float = float(p_val)
            p_display = f"{p_float:.6f}" if p_float > 0.000001 else "<0.000001"
        except (ValueError, TypeError):
            p_display = str(p_val)

        data = [
            r["Comparison"],
            w_display,
            p_display,
            r["Significant (p<0.05)"],
            round(r["Ref Mean (m)"] / 1000, 2),
            round(r["Other Mean (m)"] / 1000, 2),
        ]
        for col_idx, val in enumerate(data):
            ws2.cell(row=row, column=col_idx + 1, value=val)
            is_sig = "Yes" in str(r["Significant (p<0.05)"])
            style_body_cell(ws2, row, col_idx + 1, is_alt=is_alt, is_best=(col_idx == 3 and is_sig))
        ws2.cell(row=row, column=1).alignment = left_align

    auto_width(ws2)
    ws2.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════════
    # TABLE 2b: All-Pairwise Wilcoxon Tests
    # ══════════════════════════════════════════════════════════
    pairwise_path = "results/tables/wilcoxon_all_pairwise.csv"
    if os.path.exists(pairwise_path):
        ws2b = wb.create_sheet("Table 2b - Pairwise Wilcoxon")

        ws2b.merge_cells("A1:F1")
        ws2b.cell(row=1, column=1, value="Table 2b. All-Pairwise Wilcoxon Signed-Rank Tests")
        ws2b.cell(row=1, column=1).font = title_font
        ws2b.cell(row=1, column=1).alignment = Alignment(horizontal="left")

        ws2b.merge_cells("A2:F2")
        ws2b.cell(row=2, column=1, value="All 6 pairwise comparisons among 4 stochastic algorithms. Significance level alpha = 0.05.")
        ws2b.cell(row=2, column=1).font = caption_font

        headers2b = ["Comparison", "W Statistic", "p-value", "Significant (p<0.05)", "Mean A (km)", "Mean B (km)"]
        for i, h in enumerate(headers2b, 1):
            ws2b.cell(row=4, column=i, value=h)
        style_header_row(ws2b, 4, len(headers2b))

        df2b = pd.read_csv(pairwise_path)
        for row_idx, (_, r) in enumerate(df2b.iterrows()):
            row = row_idx + 5
            is_alt = row_idx % 2 == 1
            w_stat = r["W statistic"]
            p_val = r["p-value"]
            try:
                w_display = f"{float(w_stat):.1f}"
            except (ValueError, TypeError):
                w_display = str(w_stat)
            try:
                p_float = float(p_val)
                p_display = f"{p_float:.6f}" if p_float > 0.000001 else "<0.000001"
            except (ValueError, TypeError):
                p_display = str(p_val)

            data = [
                r["Comparison"],
                w_display,
                p_display,
                r["Significant (p<0.05)"],
                round(float(r["Mean A (m)"]) / 1000, 2),
                round(float(r["Mean B (m)"]) / 1000, 2),
            ]
            for col_idx, val in enumerate(data):
                ws2b.cell(row=row, column=col_idx + 1, value=val)
                is_sig = "Yes" in str(r["Significant (p<0.05)"])
                style_body_cell(ws2b, row, col_idx + 1, is_alt=is_alt, is_best=(col_idx == 3 and is_sig))
            ws2b.cell(row=row, column=1).alignment = left_align

        auto_width(ws2b)
        ws2b.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════════
    # TABLE 3: Parameter Sensitivity
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Table 3 - Parameters")

    ws3.merge_cells("A1:H1")
    ws3.cell(row=1, column=1, value="Table 3. MMAS Parameter Sensitivity Analysis (10 Runs per Combination)")
    ws3.cell(row=1, column=1).font = title_font
    ws3.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    ws3.merge_cells("A2:H2")
    cap_text = "27 combinations: alpha in {0.5, 1.0, 2.0}, beta in {2, 3, 5}, rho in {0.02, 0.05, 0.10}. Best row highlighted."
    ws3.cell(row=2, column=1, value=cap_text)
    ws3.cell(row=2, column=1).font = caption_font

    headers3 = ["alpha", "beta", "rho", "Mean (km)", "Std Dev (km)", "Best (km)", "Worst (km)", "Mean Time (s)"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=4, column=i, value=h)
    style_header_row(ws3, 4, len(headers3))

    df3 = pd.read_csv("results/tables/parameter_sensitivity.csv")
    df3_sorted = df3.sort_values("mean_distance")
    best_mean = df3_sorted["mean_distance"].min()

    for row_idx, (_, r) in enumerate(df3_sorted.iterrows()):
        row = row_idx + 5
        is_alt = row_idx % 2 == 1
        is_best_row = r["mean_distance"] == best_mean
        data = [
            r["alpha"], r["beta"], r["rho"],
            round(r["mean_distance"] / 1000, 2),
            round(r["std_distance"] / 1000, 2),
            round(r["best_distance"] / 1000, 2),
            round(r["worst_distance"] / 1000, 2),
            round(r["mean_time"], 4),
        ]
        for col_idx, val in enumerate(data):
            ws3.cell(row=row, column=col_idx + 1, value=val)
            style_body_cell(ws3, row, col_idx + 1, is_alt=is_alt, is_best=is_best_row)

    auto_width(ws3)
    ws3.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════════
    # TABLE 4: Scalability Analysis
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Table 4 - Scalability")

    ws4.merge_cells("A1:F1")
    ws4.cell(row=1, column=1, value="Table 4. Scalability Analysis (10 Runs per Configuration)")
    ws4.cell(row=1, column=1).font = title_font
    ws4.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    ws4.merge_cells("A2:F2")
    cap4 = "Problem sizes n = 10, 12, 15, 20 with subsets. n = 25 from main experiment."
    ws4.cell(row=2, column=1, value=cap4)
    ws4.cell(row=2, column=1).font = caption_font

    # Sub-table A: Mean Distance
    ws4.merge_cells("A4:F4")
    ws4.cell(row=4, column=1, value="(a) Mean Distance (km)")
    ws4.cell(row=4, column=1).font = subheader_font
    for c in range(1, 7):
        ws4.cell(row=4, column=c).fill = subheader_fill

    headers4a = ["n", "MMAS", "ACS", "GA", "SA", "NN"]
    for i, h in enumerate(headers4a, 1):
        ws4.cell(row=5, column=i, value=h)
    style_header_row(ws4, 5, len(headers4a))

    df4 = pd.read_csv("results/tables/scalability.csv")
    summary = pd.read_csv("results/tables/summary_table.csv")
    n_vals = sorted(df4["n_nodes"].unique())
    algos = ["MMAS", "ACS", "GA", "SA", "NN"]

    for row_idx, n in enumerate(n_vals):
        row = row_idx + 6
        is_alt = row_idx % 2 == 1
        ws4.cell(row=row, column=1, value=n)
        style_body_cell(ws4, row, 1, is_alt=is_alt)
        row_distances = {}
        for col_idx, algo in enumerate(algos, 2):
            sub = df4[(df4["n_nodes"] == n) & (df4["algorithm"] == algo)]
            if len(sub) > 0:
                val = round(sub.iloc[0]["mean_distance"] / 1000, 2)
                row_distances[col_idx] = val
                ws4.cell(row=row, column=col_idx, value=val)
            else:
                ws4.cell(row=row, column=col_idx, value="-")
            style_body_cell(ws4, row, col_idx, is_alt=is_alt)
        if row_distances:
            best_col = min(row_distances, key=row_distances.get)
            for ci in row_distances:
                style_body_cell(ws4, row, ci, is_alt=is_alt, is_best=(ci == best_col))

    # n=25 row
    row25 = len(n_vals) + 6
    is_alt25 = len(n_vals) % 2 == 1
    ws4.cell(row=row25, column=1, value=25)
    style_body_cell(ws4, row25, 1, is_alt=is_alt25)
    best_25 = 999
    best_col_25 = 2
    for col_idx, algo in enumerate(algos, 2):
        s = summary[summary["Algorithm"] == algo]
        if len(s) > 0:
            val = round(s.iloc[0]["Mean (m)"] / 1000, 2)
            ws4.cell(row=row25, column=col_idx, value=val)
            if val < best_25:
                best_25 = val
                best_col_25 = col_idx
        style_body_cell(ws4, row25, col_idx, is_alt=is_alt25)
    for ci in range(2, len(algos) + 2):
        style_body_cell(ws4, row25, ci, is_alt=is_alt25, is_best=(ci == best_col_25))

    # Sub-table B: Mean Time
    ts = row25 + 2
    ws4.merge_cells(f"A{ts}:F{ts}")
    ws4.cell(row=ts, column=1, value="(b) Mean Computation Time (seconds)")
    ws4.cell(row=ts, column=1).font = subheader_font
    for c in range(1, 7):
        ws4.cell(row=ts, column=c).fill = subheader_fill

    for i, h in enumerate(headers4a, 1):
        ws4.cell(row=ts + 1, column=i, value=h)
    style_header_row(ws4, ts + 1, len(headers4a))

    for row_idx, n in enumerate(n_vals):
        row = ts + 2 + row_idx
        is_alt = row_idx % 2 == 1
        ws4.cell(row=row, column=1, value=n)
        style_body_cell(ws4, row, 1, is_alt=is_alt)
        for col_idx, algo in enumerate(algos, 2):
            sub = df4[(df4["n_nodes"] == n) & (df4["algorithm"] == algo)]
            if len(sub) > 0:
                ws4.cell(row=row, column=col_idx, value=round(sub.iloc[0]["mean_time"], 4))
            else:
                ws4.cell(row=row, column=col_idx, value="-")
            style_body_cell(ws4, row, col_idx, is_alt=is_alt)

    row25t = ts + 2 + len(n_vals)
    ws4.cell(row=row25t, column=1, value=25)
    style_body_cell(ws4, row25t, 1, is_alt=is_alt25)
    for col_idx, algo in enumerate(algos, 2):
        s = summary[summary["Algorithm"] == algo]
        if len(s) > 0:
            ws4.cell(row=row25t, column=col_idx, value=round(s.iloc[0]["Mean Time (s)"], 4))
        style_body_cell(ws4, row25t, col_idx, is_alt=is_alt25)

    auto_width(ws4)
    ws4.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════════
    # TABLE 5: POI Details
    # ══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Table 5 - POI Details")

    ws5.merge_cells("A1:E1")
    ws5.cell(row=1, column=1, value="Table 5. Tourist Attractions in Yogyakarta Special Region")
    ws5.cell(row=1, column=1).font = title_font
    ws5.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    ws5.merge_cells("A2:E2")
    ws5.cell(row=2, column=1, value="25 representative tourist attractions across DIY Yogyakarta.")
    ws5.cell(row=2, column=1).font = caption_font

    headers5 = ["ID", "Name", "Category", "Latitude", "Longitude"]
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=4, column=i, value=h)
    style_header_row(ws5, 4, len(headers5))

    pois = pd.read_csv("data/tourist_pois.csv")
    for row_idx, (_, r) in enumerate(pois.iterrows()):
        row = row_idx + 5
        is_alt = row_idx % 2 == 1
        data = [int(r["id"]), r["name"], r["category"], round(r["latitude"], 4), round(r["longitude"], 4)]
        for col_idx, val in enumerate(data):
            ws5.cell(row=row, column=col_idx + 1, value=val)
            style_body_cell(ws5, row, col_idx + 1, is_alt=is_alt)
        ws5.cell(row=row, column=2).alignment = left_align
        ws5.cell(row=row, column=3).alignment = left_align

    auto_width(ws5, extra=4)
    ws5.sheet_view.showGridLines = False

    # ══════════════════════════════════════════════════════════
    # TABLE 6: Best Tours
    # ══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Table 6 - Best Tours")

    ws6.merge_cells("A1:C1")
    ws6.cell(row=1, column=1, value="Table 6. Best Tour Sequences per Algorithm")
    ws6.cell(row=1, column=1).font = title_font
    ws6.cell(row=1, column=1).alignment = Alignment(horizontal="left")

    ws6.merge_cells("A2:C2")
    ws6.cell(row=2, column=1, value="Tour as 1-indexed POI IDs forming a Hamiltonian cycle.")
    ws6.cell(row=2, column=1).font = caption_font

    headers6 = ["Algorithm", "Distance (km)", "Tour Sequence"]
    for i, h in enumerate(headers6, 1):
        ws6.cell(row=4, column=i, value=h)
    style_header_row(ws6, 4, len(headers6))

    tours = pd.read_csv("results/tables/tours.csv")
    best_tours = tours.loc[tours.groupby("algorithm")["distance_m"].idxmin()]
    for row_idx, (_, r) in enumerate(best_tours.iterrows()):
        row = row_idx + 5
        is_alt = row_idx % 2 == 1
        data = [r["algorithm"], round(r["distance_m"] / 1000, 2), r["tour"]]
        for col_idx, val in enumerate(data):
            ws6.cell(row=row, column=col_idx + 1, value=val)
            style_body_cell(ws6, row, col_idx + 1, is_alt=is_alt)
        ws6.cell(row=row, column=1).alignment = left_align
        ws6.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    auto_width(ws6, extra=4)
    ws6.column_dimensions["C"].width = 80
    ws6.sheet_view.showGridLines = False

    # ── Save ─────────────────────────────────────────────────
    out_path = os.path.join("results", "tables", "journal_tables.xlsx")
    wb.save(out_path)
    print(f"Excel tables saved to: {out_path}")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    create_journal_tables()
